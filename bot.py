from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    filters,
)
from datetime import datetime, time
from openpyxl import Workbook, load_workbook
import os

TOKEN = "8556753858:AAGjgVU7rElIgMhMOIMktD_mb9OdlbtZV-o"
ADMIN_PASSWORD = "2025cia"

BREADS = [
    "Ağ çiabatta",
    "Boz çiabatta",
    "Kömürlü çörək",
    "Kömürlü bol toxumlu",
    "Zeytunlu Çörək",
    "Boz batar",
    "Alman Çörəyi",
    "Bol toxumlu",
    "Baget",
    "Ağ batar",
]

orders = {}
waiting_quantity = {}
subscribers = set()

def today_file():
    return f"orders_{datetime.now().date()}.xlsx"

def main_keyboard():
    return ReplyKeyboardMarkup([["🟢 Sabah üçün sifariş"]], resize_keyboard=True)

def bread_keyboard():
    buttons = [[b] for b in BREADS]
    buttons.append(["✅ Sifarişi bitir"])
    return ReplyKeyboardMarkup(buttons, resize_keyboard=True)

def phone_keyboard():
    return ReplyKeyboardMarkup(
        [[KeyboardButton("📞 Telefonu göndər", request_contact=True)]],
        resize_keyboard=True,
        one_time_keyboard=True
    )

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    subscribers.add(update.message.from_user.id)
    await update.message.reply_text(
        "Xoş gəldiniz! 🥖\nSabah üçün sifariş edə bilərsiniz.",
        reply_markup=main_keyboard()
    )

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.message.from_user.id
    text = update.message.text

    if text == "🟢 Sabah üçün sifariş":
        orders[uid] = {}
        await update.message.reply_text("Çörəyi seçin:", reply_markup=bread_keyboard())
        return

    if text in BREADS:
        waiting_quantity[uid] = text
        await update.message.reply_text(f"{text} üçün say yazın:")
        return

    if text == "✅ Sifarişi bitir":
        if not orders.get(uid):
            await update.message.reply_text("Sifariş boşdur.")
            return
        await update.message.reply_text("Telefon nömrənizi göndərin:", reply_markup=phone_keyboard())
        return

    if uid in waiting_quantity:
        if not text.isdigit():
            await update.message.reply_text("Rəqəm yazın.")
            return
        bread = waiting_quantity.pop(uid)
        orders.setdefault(uid, {})
        orders[uid][bread] = orders[uid].get(bread, 0) + int(text)
        await update.message.reply_text(f"Əlavə olundu: {bread}")

async def handle_contact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.message.from_user.id
    phone = update.message.contact.phone_number
    date = str(datetime.now().date())
    file = today_file()

    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(["Tarix", "Telefon", "Çörək", "Say"])
        wb.save(file)

    wb = load_workbook(file)
    ws = wb.active

    for bread, qty in orders.get(uid, {}).items():
        ws.append([date, phone, bread, qty])

    wb.save(file)
    orders[uid] = {}

    await update.message.reply_text("✅ Sifariş saxlanıldı!", reply_markup=main_keyboard())

async def admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args or context.args[0] != ADMIN_PASSWORD:
        await update.message.reply_text("❌ Yanlış parol.")
        return

    file = today_file()
    if not os.path.exists(file):
        await update.message.reply_text("Bu gün sifariş yoxdur.")
        return

    await update.message.reply_text(f"📊 Bugünkü Excel faylı hazırdır:\n{file}")

async def clear_orders(context: ContextTypes.DEFAULT_TYPE):
    orders.clear()

async def daily_notify(context: ContextTypes.DEFAULT_TYPE):
    for uid in subscribers:
        try:
            await context.bot.send_message(
                chat_id=uid,
                text="🥖 Sabah üçün təzə çörək sifariş etmək istərdiniz?"
            )
        except:
            pass

def main():
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("admin", admin))
    app.add_handler(MessageHandler(filters.CONTACT, handle_contact))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    app.job_queue.run_daily(clear_orders, time(hour=6, minute=0))
    app.job_queue.run_daily(daily_notify, time(hour=18, minute=0))

    print("Bot işə düşdü...")
    app.run_polling()

if __name__ == "__main__":
    main()
